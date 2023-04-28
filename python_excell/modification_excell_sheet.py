from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
	wb = load_workbook("filename")
	print(wb.sheetnames)
	#ws = work Sheet
	ws = wb["Sheet1"]
	print(ws)
	cell = ws.cell(1,1)
	print(cell.value)
	print(f"max of rows it's: {ws.max_row}")
	
	for row in range(2, ws.max_row + 1):
	#	print(row)
		cell = ws.cell(row,3)
	#	print(cell.value)
		correted_price = cell.value * 0.9
	#	print(correted_price)
		#how to add a values to new column
		corrected_price_cell = ws.cell(row,4)
		corrected_price_cell.value = correted_price
	ws.cell(1,ws.max_row).value = "corrected price"
	
	values = Reference(ws,
			  min_row = 2,
			  max_row = ws.max_row,
			  min_col = 4,
			  max_col = 4)
	
	chart = BarChart()
	chart.add_data(values)
	ws.add_chart(chart, "e2")
	
	wb.save("filename")
