from openpyxl import Workbook, load_workbook

wb = load_workbook('transactions.xlsx')

ws = wb.active
print(wb.sheetnames)
print(wb['Sheet1'])
#create new sheet
wb.create_sheet('Test')

print(wb.sheetnames)
