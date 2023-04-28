from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#creation of new excell workbook
wb = load_workbook("kitty.xlsx")
ws= wb.active
print(ws)

for row in range(1, 11):
    for col in range(1,5):
        char = get_column_letter(col)
#to fill the fields by theirs names:
        ws[char + str(row)] = char + str(row)

wb.save('kitty.xlsx')
