from openpyxl import Workbook, load_workbook

'''test how it works:
for col in range(0, 4):
    char = chr(65 + col)
    print(char)
'''

#creation of new excell workbook
wb = load_workbook("kitty.xlsx")
ws= wb.active
print(ws)

for row in range(1, 11):
    for col in range(0,4):
        char = chr(65 + col)
        ws[char + str(row)] = char + str(row + 2)

wb.save("kitty.xlsx")
