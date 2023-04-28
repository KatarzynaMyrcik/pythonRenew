from openpyxl import Workbook, load_workbook

wb = load_workbook('transactions.xlsx')

ws = wb.active
print(ws['a1'].value)

ws['a1'].value = "Transaction_id"
print(ws['a1'].value)

wb.save('tranzakcje.xlsx')
