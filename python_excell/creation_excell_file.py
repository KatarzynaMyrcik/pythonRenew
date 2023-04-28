from openpyxl import Workbook, load_workbook

#creation of new excell workbook
wb = Workbook()
#creation of new sheet
ws = wb.active
print(wb.sheetnames)
#renaming of new sheet
ws.title = "Data"
print(wb.sheetnames)

for i in range(3):
    ws.append(['Kity', 'is', 'the', 'best', '!', 10])
ws.append(['end'])

wb.save("kitty.xlsx")
