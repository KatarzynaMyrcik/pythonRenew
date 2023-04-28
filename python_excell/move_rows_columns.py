from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#creation of new excell workbook
wb = load_workbook("kitty.xlsx")
ws = wb.active

ws.move_range("D1:E12", rows=2, cols=1,)

wb.save("kitty.xlsx")
