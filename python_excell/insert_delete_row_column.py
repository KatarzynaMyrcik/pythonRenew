from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#creation of new excell workbook
wb = load_workbook("kitty.xlsx")
ws = wb.active


#ws.insert_cols(3)
#ws.insert_rows(6)
ws.delete_cols(3)
ws.delete_rows(6)
wb.save("kitty.xlsx")
